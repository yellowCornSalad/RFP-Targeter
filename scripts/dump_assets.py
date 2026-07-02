"""특허·인증 전체 덤프 → 분류 + JSON 산출."""
import json
from pathlib import Path

import openpyxl

OUT = Path(r"D:\RFP-Targeter\data\raw\company_assets.json")


def read_patents(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = {"domestic_registered": [], "foreign_registered": [], "domestic_pending": [], "foreign_pending": []}

    # 등록현황 시트
    ws = wb["등록현황"]
    cur = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        head = cells[0]
        if "(국내) 등록특허" in head:
            cur = "domestic_registered"; continue
        if "(해외) 등록특허" in head:
            cur = "foreign_registered"; continue
        if head in ("NO", "no", "No") or cells[1] == "무형자산 등록코드":
            continue
        if cur and head.isdigit():
            rows[cur].append({
                "no": head,
                "app_no": cells[2],
                "app_date": cells[3],
                "reg_no": cells[4],
                "reg_date": cells[5],
                "owner": cells[6],
                "title": cells[7],
                "status": cells[8],
                "expire": cells[9],
                "note": cells[10],
            })

    # 출원현황 시트
    ws = wb["출원현황"]
    cur = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        head = cells[0]
        if "(국내) 출원특허" in head:
            cur = "domestic_pending"; continue
        if "(해외) 출원특허" in head:
            cur = "foreign_pending"; continue
        if head in ("NO", "no", "No") or cells[1] == "구분":
            continue
        if cur and head.isdigit():
            rows[cur].append({
                "no": head,
                "type": cells[1],
                "app_no": cells[2],
                "app_date": cells[3],
                "status": cells[4],
                "title": cells[5],
            })
    return rows


def read_certs(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        rows.append({
            "기관": cells[0], "유형": cells[1], "기술명": cells[2], "보유자": cells[3],
        })
    return rows


patents = read_patents(r"D:\RFP-Targeter\data\raw\enki_ip.xlsx")
certs = read_certs(r"D:\RFP-Targeter\data\raw\enki_certs.xlsx")

data = {"patents": patents, "certs": certs}
OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"국내 등록: {len(patents['domestic_registered'])}건")
print(f"해외 등록: {len(patents['foreign_registered'])}건")
print(f"국내 출원: {len(patents['domestic_pending'])}건")
print(f"해외 출원: {len(patents['foreign_pending'])}건")
print(f"인증/SW: {len(certs)}건")
print(f"저장 → {OUT}")
